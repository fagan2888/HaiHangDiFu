# _*_ coding: utf-8 _*_

import locale

from datetime import datetime

from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl import worksheet

from select_db import HHDF
from update import update


def main():
    update()

    # 设置本地语言符号显示为中文
    locale.setlocale(locale.LC_CTYPE, 'chinese')

    book = Workbook()
    sheet = book.active
    sheet.title = "业务统计表"

    # 设置表标题
    sheet.merge_cells("A1:G1")
    sheet["A1"].font = Font(name='微软雅黑', size=14, bold=True)
    sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["A1"].value = "云南分公司海航地服合作项目情况汇总表"

    # 设置统计时间范围
    sheet.merge_cells("A2:G2")
    sheet["A2"].font = Font(name='微软雅黑', size=10)
    sheet["A2"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["A2"].value = f"统计时间2019年01月01日 至 \
{datetime.now().strftime('%m月%d日')}"

    hh = HHDF()

    # 设置加粗字体样式
    bold_style = NamedStyle(name='bold_style')
    bold_style.font = Font(name='微软雅黑', size=12, bold=True)
    bold_style.alignment = Alignment(horizontal='center', vertical='center')
    side = Side(border_style='thin', color='000000')
    bold_style.border = Border(left=side, right=side, top=side, bottom=side)
    book.add_named_style(bold_style)

    # 设置不加粗字体样式,用于显示数据
    db_style = NamedStyle(name='db_style')
    db_style.font = Font(name='微软雅黑', size=12)
    db_style.alignment = Alignment(horizontal='center', vertical='center')
    side = Side(border_style='thin', color='000000')
    db_style.border = Border(left=side, right=side, top=side, bottom=side)
    book.add_named_style(db_style)

    # 设置表格中的数据内容
    rows = []

    tb_head = ("销售产品",
               "类别",
               f"{hh.three_days_ago[5:7]}月{hh.three_days_ago[8:10]}日",
               f"{hh.two_days_ago[5:7]}月{hh.two_days_ago[8:10]}日",
               f"{hh.one_days_ago[5:7]}月{hh.one_days_ago[8:10]}日",
               "月度累计",
               "年度累计")
    rows.append(tb_head)

    row = ("交通意外保险",
           "件数",
           hh.three_days_ago_jiao_tong_count,
           hh.two_days_ago_jiao_tong_count,
           hh.one_days_ago_jiao_tong_count,
           hh.month_jiao_tong_count,
           hh.year_jiao_tong_count)
    rows.append(row)

    row = ("交通意外保险",
           "保费",
           hh.three_days_ago_jiao_tong_sum,
           hh.two_days_ago_jiao_tong_sum,
           hh.one_days_ago_jiao_tong_sum,
           hh.month_jiao_tong_sum,
           hh.year_jiao_tong_sum)
    rows.append(row)

    row = ("航空意外险",
           "件数",
           hh.three_days_ago_hang_kong_count,
           hh.two_days_ago_hang_kong_count,
           hh.one_days_ago_hang_kong_count,
           hh.month_hang_kong_count,
           hh.year_hang_kong_count)
    rows.append(row)

    row = ("航空意外险",
           "保费",
           hh.three_days_ago_hang_kong_sum,
           hh.two_days_ago_hang_kong_sum,
           hh.one_days_ago_hang_kong_sum,
           hh.month_hang_kong_sum,
           hh.year_hang_kong_sum)
    rows.append(row)

    row = ("合计",
           "件数",
           hh.three_days_ago_count,
           hh.two_days_ago_count,
           hh.one_days_ago_count,
           hh.month_count,
           hh.year_count)
    rows.append(row)

    row = ("合计",
           "保费",
           hh.three_days_ago_sum,
           hh.two_days_ago_sum,
           hh.one_days_ago_sum,
           hh.month_sum,
           hh.year_sum)
    rows.append(row)

    # 将列表中的数据写入表格
    nrow = 3
    while nrow <= 9:
        ncol = 1
        while ncol <= 7:
            sheet.cell(row=nrow, column=ncol).value = rows[nrow - 3][ncol - 1]
            if ncol > 2 and nrow > 3:
                sheet.cell(row=nrow, column=ncol).style = db_style
            else:
                sheet.cell(row=nrow, column=ncol).style = bold_style
            ncol += 1
        sheet.row_dimensions[nrow].height = 26
        nrow += 1

    # 合并险种部分的单元格
    sheet.merge_cells("A4:A5")
    sheet.merge_cells("A6:A7")
    sheet.merge_cells("A8:A9")

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10

    sheet['A11'].value = '注：当日统计时间为早10点至次日早10点（24小时），与班组工作时间一致'
    sheet.merge_cells("A11:G11")

    book.save("海航地服业务统计表.xlsx")


if __name__ == '__main__':
    main()
